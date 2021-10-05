import openpyxl
import yaml
import json

from openpyxl import load_workbook

from openpyxl.utils import get_column_letter

def find_last_items_row(ws, start):
    i = start
    while i < ws.max_row:
        val = ws[f'E{i}'].value
        if not val:
            return i
        i += 1
    return ws.max_row + 1

def find_last_ref_col(ws, start):
    i = start
    while i < ws.max_column:
        # val = ws.cell(3, start).value
        val = ws.cell(3, i).value
        if not val:
            return i
        i += 1
    return ws.max_column

def col_next_cell(ws, col, startRow):
    r = startRow
    while r < ws.max_row:
        val = ws.cell(r, col).value
        if val:
            return r
        r += 1
    return -1


def append_subcategories(ws, cIndex, subcategories):
    eIndex = cIndex + 1
    subcatId = ws[f'D{cIndex}'].value
    subcatName = ws[f'C{cIndex}'].value
    subcategory = { "id": subcatId, "name": subcatName, "items": [] }
    subcategories.append(subcategory)
    last_items_row = find_last_items_row(ws, eIndex)

    #print (subcatName)
    #print (last_items_row)

    # Assuming same ref colums for all items (possibly empty)
    firstRefCol = 18 + 7
    lastRefCol = find_last_ref_col(ws, firstRefCol)

    for i in range(eIndex, last_items_row):
        itemId = ws[f'F{eIndex}'].value
        itemName = ws[f'E{eIndex}'].value
        itemDescr = ws[f'G{eIndex}'].value
        menuItem = { "id": itemId, "name": itemName, "description": itemDescr, "allergens": [], "infos": [], "sizes": [], "ingredients": []}
        subcategory["items"].append(menuItem)
        eIndex += 1 # ?

        #print(itemName)

        for k in range(0, 2):
            allergIndex = 8 + k
            allergValue = ws.cell(i, allergIndex).value
            if allergValue:
                allergName = ws.cell(4, allergIndex).value
                menuItem["infos"].append(allergName)

        for k in range(3, 17):
            allergIndex = 8 + k
            allergValue = ws.cell(i, allergIndex).value
            if allergValue:
                allergName = ws.cell(4, allergIndex).value
                menuItem["allergens"].append(allergName)

        # print(openpyxl.utils.cell.get_column_letter(lastRefCol))

        for k in range(firstRefCol, lastRefCol, 3):
            ref = {
                "ref": ws.cell(i, k).value,
                "label": ws.cell(i, k + 1).value,
                "price": ws.cell(i, k + 2).value
            }

            if ref["ref"] or ref["label"] or ref["price"]:
                menuItem["sizes"].append(ref)

        if lastRefCol < ws.max_column:
            for k in range(lastRefCol, ws.max_column + 1):
                ingVal = ws.cell(i, k).value
                if ingVal:
                    ingLabel = ws.cell(4, k).value
                    menuItem["ingredients"].append(ingLabel)


wb = load_workbook(filename = 'DATABASE_MENU_VENDITA.xlsx', data_only=True)

result = { "categories": [] }

for sheet in wb.sheetnames:

    ws = wb[sheet]

    catId = ws["B3"].value
    catName = ws["A3"].value
    catPop = ws["A4"].value
  
    category = { "id": catId, "name": catName, "subcategories": [], "ingredients": [], "pop": catPop }
    result["categories"].append(category)

    # Assuming same ref colums for all items (possibly empty)
    firstRefCol = 18 + 7
    lastRefCol = find_last_ref_col(ws, firstRefCol)
    if lastRefCol < ws.max_column:
        for k in range(lastRefCol, ws.max_column + 1):
            ingLabel = ws.cell(4, k).value
            category["ingredients"].append(ingLabel)

    cIndex = 4

    append_subcategories(ws, cIndex, category["subcategories"])

    while True:
        cIndex = col_next_cell(ws, 3, cIndex + 1)
        if cIndex > -1:
            append_subcategories(ws, cIndex, category["subcategories"])
        else:
            break

yamlResult = yaml.dump(result)
jsonResult = json.dumps(result)
# print (yamlResult)

f = open("result.yml", "w")
f.write(yamlResult)
f.close()

f = open("result.json", "w")
f.write(jsonResult)
f.close()