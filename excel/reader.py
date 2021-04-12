def find_start_row(ws):
  i = 1
  while  i < ws.max_row:
    val = ws[f'A{i}'].value
    if val:
      return i
    i += 1
  return ws.max_row

def find_end_row(ws, start):
  i = start
  while  i < ws.max_row:
    val = ws[f'A{i}'].value
    if not val:
      return i
    i += 2
  return ws.max_row

def find_end_col(ws, start):
  i = start
  while  i < ws.max_column:
    val = ws.cell(2,i).value
    if not val:
      return i
    i += 1
    return ws.max_column
##

START_COL = 4

import openpyxl
from openpyxl import load_workbook
from columnar import columnar
wb = load_workbook(filename = 'RicettePub.xlsx', data_only=True)
ws = wb['Sheet2']
start_row = find_start_row(ws)
end_row = find_end_row(ws, start_row)
end_col = find_end_col(ws, START_COL)
#print(openpyxl.utils.cell.get_column_letter(end_col))

for r in range(start_row, end_row, 2):
  val = ws[f'B{r}'].value
  price =  ws[f'C{r+1}'].value
  print()
  print(f'{val} € {price}')
  print()
  data = []
  for cell in ws.iter_cols(min_row=2, max_row=2, min_col=START_COL, max_col=end_col):
    #print(f'--{cell[0].value}')
    #print(f'--{cell[0].column}')
    if (ws.cell(r,cell[0].column).value):
      #print(ws.cell(r,cell[0].column).value)
      #print(f'--{cell[0].value} Qta {ws.cell(r,cell[0].column).value} € {ws.cell(r + 1,cell[0].column).value} (% {ws.cell(r + 1,cell[0].column).value / price})') 
      data.append([f'{cell[0].value}', f'Qta {ws.cell(r,cell[0].column).value}', f'€ {ws.cell(r + 1,cell[0].column).value}', f'% {ws.cell(r + 1,cell[0].column).value / price}'])

  #table = columnar(data, headers=['Race', 'Date', 'Location', 'Distance'], patterns=patterns)
  table = columnar(data)
  print(table)
  



